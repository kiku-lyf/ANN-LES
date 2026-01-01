import sys
import os
import time
from mpi4py import MPI
import numpy as np
from numba import njit
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

# MPI初始化
comm = MPI.COMM_WORLD
rank = comm.Get_rank()
size = comm.Get_size()

# 创建一个空的DataFrame（只在rank 0）
if rank == 0:
    df = pd.DataFrame()
    df.to_excel('.train.xlsx', index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

Q = 9
NX = 1023
NY = 1023
U = 0.1

e = np.array([[0, 0], [1, 0], [0, 1], [-1, 0], [0, -1], [1, 1], [-1, 1], [-1, -1], [1, -1]])
w = np.array([4.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 36, 1.0 / 36, 1.0 / 36, 1.0 / 36])

dx = 1.0
dy = 1.0
Lx = dx * NY
Ly = dy * NX
dt = dx
c = dx / dt
rho0 = 1.0
Re = 10000
niu = U * Lx / Re
tau_f = 3.0 * niu + 0.5

# MPI域分解：在y方向分割
Lx_mpi = NX + 1
Ly_mpi = int((NY + 1) / size)  # 每个进程的y方向大小
LYv = Ly_mpi + 2  # 添加两个虚格点用于通信

# MPI通信缓冲区
allsend = np.zeros(2 * (Q + 2 + 1) * Lx_mpi, dtype=np.float64)
allReceive = np.zeros(2 * (Q + 2 + 1) * Lx_mpi, dtype=np.float64)

# MPI局部数组
mpi_f = np.zeros((Lx_mpi, LYv, Q), dtype=np.float64)
mpi_u = np.zeros((Lx_mpi, LYv, 2), dtype=np.float64)
mpi_rho = np.ones((Lx_mpi, LYv), dtype=np.float64)
mpi_F = np.zeros((Lx_mpi, LYv, Q), dtype=np.float64)
mpi_u0 = np.zeros((Lx_mpi, LYv, 2), dtype=np.float64)

# 全局数组（用于收集完整数据）
if rank == 0:
    u_global = np.zeros((NX + 1, NY + 1, 2), dtype=np.float64)
else:
    u_global = None


@njit
def tforce(T,u,Ub):
    # 8×8盒式滤波，计算滤波后的速度和亚格子应力
    # 边界点权重0.5，内部点权重1.0，角点权重0.25
    for i in range(1,127):
        for j in range(1,127):
            ujj=0.0
            uii = 0.0
            uij=0.0
            ui=0.0
            uj=0.0
            sumw = 0.0  # 权重和
            
            for i1 in range(8):
                for j1 in range(8):
                    # 确定权重：边界0.5，角点0.25，内部1.0
                    if (i1 == 0 or i1 == 7) and (j1 == 0 or j1 == 7):
                        wgt = 0.25  # 角点
                    elif i1 == 0 or i1 == 7 or j1 == 0 or j1 == 7:
                        wgt = 0.5   # 边界点
                    else:
                        wgt = 1.0   # 内部点
                    
                    ux = u[i*8+i1][j*8+j1][0]
                    uy = u[i*8+i1][j*8+j1][1]
                    
                    ui += ux * wgt
                    uj += uy * wgt
                    uij += ux * uy * wgt
                    uii += ux * ux * wgt
                    ujj += uy * uy * wgt
                    sumw += wgt
            
            # 归一化（使用实际权重和）
            ui_norm = ui / sumw
            uj_norm = uj / sumw
            uij_norm = uij / sumw
            uii_norm = uii / sumw
            ujj_norm = ujj / sumw
            
            # 计算亚格子应力：τ_ij = overline(u_i u_j) - ū_i ū_j
            T[i][j][0] = uii_norm - ui_norm * ui_norm   # txx
            T[i][j][1] = uij_norm - ui_norm * uj_norm   # txy
            T[i][j][2] = uij_norm - ui_norm * uj_norm   # tyx (对称)
            T[i][j][3] = ujj_norm - uj_norm * uj_norm   # tyy
            Ub[i][j][0] = ui_norm                       # ux~
            Ub[i][j][1] = uj_norm                       # uy~

@njit
def ronudU(Ub,A,Am):
    #梯度张量 A_ij = ∂_i u_j
    # 使用中心差分，边界使用单边差分
    for i in range(1,127):
        for j in range(1,127):
            # 内部点使用中心差分
            if i < 126 and j < 126:
                # Axx = ∂_x u_x = (u_x[i][j+1] - u_x[i][j-1]) / 2
                # 简化为中心差分：Axx = u_x[i][j+1] - u_x[i][j]
                A[i][j][0] = Ub[i][j + 1][0] - Ub[i][j][0]  # Axx = ∂_x u_x
                A[i][j][1] = Ub[i][j + 1][1] - Ub[i][j][1]  # Axy = ∂_x u_y
                A[i][j][2] = Ub[i + 1][j][0] - Ub[i][j][0]  # Ayx = ∂_y u_x
                A[i][j][3] = Ub[i + 1][j][1] - Ub[i][j][1]  # Ayy = ∂_y u_y
            # 边界点使用单边差分
            elif i == 126 or j == 126:
                if j == 126:
                    A[i][j][0] = Ub[i][j][0] - Ub[i][j-1][0]  # Axx
                    A[i][j][1] = Ub[i][j][1] - Ub[i][j-1][1]  # Axy
                else:
                    A[i][j][0] = Ub[i][j + 1][0] - Ub[i][j][0]  # Axx
                    A[i][j][1] = Ub[i][j + 1][1] - Ub[i][j][1]  # Axy
                if i == 126:
                    A[i][j][2] = Ub[i][j][0] - Ub[i-1][j][0]  # Ayx
                    A[i][j][3] = Ub[i][j][1] - Ub[i-1][j][1]  # Ayy
                else:
                    A[i][j][2] = Ub[i + 1][j][0] - Ub[i][j][0]  # Ayx
                    A[i][j][3] = Ub[i + 1][j][1] - Ub[i][j][1]  # Ayy
            # 计算梯度张量的模
            Am[i][j]=np.sqrt(A[i][j][0]*A[i][j][0]+A[i][j][1]*A[i][j][1]+A[i][j][2]*A[i][j][2]+A[i][j][3]*A[i][j][3])



@njit
def feq(k, rho, u):
    eu = e[k][0] * u[0] + e[k][1] * u[1]
    uv = u[0] * u[0] + u[1] * u[1]
    return w[k] * rho * (1.0 + 3.0 * eu + 4.5 * eu ** 2 - 1.5 * uv)


@njit
def init_mpi(mpi_f, mpi_u, mpi_rho, rank, size, Ly_mpi, LYv, Lx_mpi):
    """初始化MPI局部数组"""
    me_y = rank
    MY = size - 1
    
    # 初始化局部数组
    for i in range(Lx_mpi):
        for j in range(1, LYv - 1):
            mpi_rho[i][j] = 1.0
            mpi_u[i][j][0] = 0.0
            mpi_u[i][j][1] = 0.0
            for k in range(Q):
                mpi_f[i][j][k] = feq(k, mpi_rho[i][j], mpi_u[i][j])
    
    # 上边界（顶盖驱动流）
    if me_y == MY:
        for i in range(Lx_mpi):
            mpi_u[i][LYv - 2][0] = U
    
    return me_y, MY


@njit
def befor_send(allsend, mpi_f, mpi_u, mpi_rho, Lx_mpi, Q, LYv):
    """准备发送的边界数据"""
    for i in range(Lx_mpi):
        for k in range(Q):
            # 上边界数据
            allsend[i * Q + k] = mpi_f[i][LYv - 2][k]
            # 下边界数据
            allsend[Lx_mpi * Q + 3 * Lx_mpi + i * Q + k] = mpi_f[i][1][k]
        
        for m in range(2):
            # 上边界速度
            allsend[Lx_mpi * Q + i * 2 + m] = mpi_u[i][LYv - 2][m]
            # 下边界速度
            allsend[2 * Lx_mpi * Q + 3 * Lx_mpi + i * 2 + m] = mpi_u[i][1][m]
        
        # 上边界密度
        allsend[Lx_mpi * Q + 2 * Lx_mpi + i] = mpi_rho[i][LYv - 2]
        # 下边界密度
        allsend[2 * Lx_mpi * Q + 5 * Lx_mpi + i] = mpi_rho[i][1]


def mpi_sendarecv(me_y, MY):
    """MPI通信：发送和接收边界数据"""
    # 如果不是上边界和下边界
    if me_y != 0 and me_y != MY:
        # 向上发送，从下面接收
        allReceive[Lx_mpi * Q + 3 * Lx_mpi:] = comm.sendrecv(
            sendobj=allsend[:Lx_mpi * Q + 3 * Lx_mpi], 
            dest=me_y + 1, sendtag=1,
            recvbuf=None, source=me_y - 1, recvtag=1
        )
        # 向下发送，从上面接收
        allReceive[:Lx_mpi * Q + 3 * Lx_mpi] = comm.sendrecv(
            sendobj=allsend[Lx_mpi * Q + 3 * Lx_mpi:], 
            dest=me_y - 1, sendtag=2,
            recvbuf=None, source=me_y + 1, recvtag=2
        )
    # 下边界（rank 0）
    elif me_y == 0:
        allReceive[:Lx_mpi * Q + 3 * Lx_mpi] = comm.sendrecv(
            sendobj=allsend[:Lx_mpi * Q + 3 * Lx_mpi], 
            dest=me_y + 1, sendtag=1,
            recvbuf=None, source=me_y + 1, recvtag=2
        )
    # 上边界（最后一个rank）
    else:
        allReceive[Lx_mpi * Q + 3 * Lx_mpi:] = comm.sendrecv(
            sendobj=allsend[Lx_mpi * Q + 3 * Lx_mpi:], 
            dest=me_y - 1, sendtag=2,
            recvbuf=None, source=me_y - 1, recvtag=1
        )


@njit
def after_receive(mpi_f, mpi_u, mpi_rho, allReceive, Lx_mpi, Q, LYv):
    """接收边界数据并更新虚格点"""
    for i in range(Lx_mpi):
        for k in range(Q):
            # 下边界虚格点（从下面进程接收）
            mpi_f[i][0][k] = allReceive[Lx_mpi * Q + 3 * Lx_mpi + i * Q + k]
            # 上边界虚格点（从上面进程接收）
            mpi_f[i][LYv - 1][k] = allReceive[i * Q + k]
        
        for m in range(2):
            mpi_u[i][0][m] = allReceive[2 * Lx_mpi * Q + 3 * Lx_mpi + i * 2 + m]
            mpi_u[i][LYv - 1][m] = allReceive[Lx_mpi * Q + i * 2 + m]
        
        mpi_rho[i][0] = allReceive[2 * Lx_mpi * Q + 5 * Lx_mpi + i]
        mpi_rho[i][LYv - 1] = allReceive[Lx_mpi * Q + 2 * Lx_mpi + i]


@njit
def mpi_evolution(mpi_f, mpi_u, mpi_rho, mpi_F, mpi_u0, me_y, MY, Lx_mpi, LYv):
    """MPI版本的LBM演化"""
    # 第一步：碰撞和迁移
    if me_y != 0 and me_y != MY:
        # 内部进程
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 1):
                for k in range(Q):
                    ip = i - e[k][0]
                    jp = j - e[k][1]
                    mpi_F[i][j][k] = mpi_f[ip][jp][k] + (feq(k, mpi_rho[ip][jp], mpi_u[ip][jp]) - mpi_f[ip][jp][k]) / tau_f
        
        # 更新宏观量
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 1):
                mpi_u0[i][j][0] = mpi_u[i][j][0]
                mpi_u0[i][j][1] = mpi_u[i][j][1]
                mpi_rho[i][j] = 0
                mpi_u[i][j][0] = 0
                mpi_u[i][j][1] = 0
                for k in range(Q):
                    mpi_f[i][j][k] = mpi_F[i][j][k]
                    mpi_rho[i][j] += mpi_f[i][j][k]
                    mpi_u[i][j][0] += e[k][0] * mpi_f[i][j][k]
                    mpi_u[i][j][1] += e[k][1] * mpi_f[i][j][k]
                mpi_u[i][j][0] /= mpi_rho[i][j]
                mpi_u[i][j][1] /= mpi_rho[i][j]
        
        # 左右边界
        for j in range(1, LYv - 1):
            for k in range(Q):
                mpi_rho[0][j] = mpi_rho[1][j]
                mpi_f[0][j][k] = feq(k, mpi_rho[0][j], mpi_u[0][j]) + mpi_f[1][j][k] - feq(k, mpi_rho[1][j], mpi_u[1][j])
                mpi_rho[Lx_mpi - 1][j] = mpi_rho[Lx_mpi - 2][j]
                mpi_f[Lx_mpi - 1][j][k] = feq(k, mpi_rho[Lx_mpi - 1][j], mpi_u[Lx_mpi - 1][j]) + mpi_f[Lx_mpi - 2][j][k] - feq(k, mpi_rho[Lx_mpi - 2][j], mpi_u[Lx_mpi - 2][j])
    
    # 下边界进程（rank 0）
    elif me_y == 0:
        for i in range(1, Lx_mpi - 1):
            for j in range(2, LYv - 1):
                for k in range(Q):
                    ip = i - e[k][0]
                    jp = j - e[k][1]
                    mpi_F[i][j][k] = mpi_f[ip][jp][k] + (feq(k, mpi_rho[ip][jp], mpi_u[ip][jp]) - mpi_f[ip][jp][k]) / tau_f
        
        for i in range(1, Lx_mpi - 1):
            for j in range(2, LYv - 1):
                mpi_u0[i][j][0] = mpi_u[i][j][0]
                mpi_u0[i][j][1] = mpi_u[i][j][1]
                mpi_rho[i][j] = 0
                mpi_u[i][j][0] = 0
                mpi_u[i][j][1] = 0
                for k in range(Q):
                    mpi_f[i][j][k] = mpi_F[i][j][k]
                    mpi_rho[i][j] += mpi_f[i][j][k]
                    mpi_u[i][j][0] += e[k][0] * mpi_f[i][j][k]
                    mpi_u[i][j][1] += e[k][1] * mpi_f[i][j][k]
                mpi_u[i][j][0] /= mpi_rho[i][j]
                mpi_u[i][j][1] /= mpi_rho[i][j]
        
        # 左右边界
        for j in range(2, LYv - 1):
            for k in range(Q):
                mpi_rho[0][j] = mpi_rho[1][j]
                mpi_f[0][j][k] = feq(k, mpi_rho[0][j], mpi_u[0][j]) + mpi_f[1][j][k] - feq(k, mpi_rho[1][j], mpi_u[1][j])
                mpi_rho[Lx_mpi - 1][j] = mpi_rho[Lx_mpi - 2][j]
                mpi_f[Lx_mpi - 1][j][k] = feq(k, mpi_rho[Lx_mpi - 1][j], mpi_u[Lx_mpi - 1][j]) + mpi_f[Lx_mpi - 2][j][k] - feq(k, mpi_rho[Lx_mpi - 2][j], mpi_u[Lx_mpi - 2][j])
        
        # 下边界（物理边界）
        for i in range(Lx_mpi):
            for k in range(Q):
                mpi_rho[i][1] = mpi_rho[i][2]
                mpi_f[i][1][k] = feq(k, mpi_rho[i][1], mpi_u[i][1]) + mpi_f[i][2][k] - feq(k, mpi_rho[i][2], mpi_u[i][2])
    
    # 上边界进程（最后一个rank）
    else:
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 2):
                for k in range(Q):
                    ip = i - e[k][0]
                    jp = j - e[k][1]
                    mpi_F[i][j][k] = mpi_f[ip][jp][k] + (feq(k, mpi_rho[ip][jp], mpi_u[ip][jp]) - mpi_f[ip][jp][k]) / tau_f
        
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 2):
                mpi_u0[i][j][0] = mpi_u[i][j][0]
                mpi_u0[i][j][1] = mpi_u[i][j][1]
                mpi_rho[i][j] = 0
                mpi_u[i][j][0] = 0
                mpi_u[i][j][1] = 0
                for k in range(Q):
                    mpi_f[i][j][k] = mpi_F[i][j][k]
                    mpi_rho[i][j] += mpi_f[i][j][k]
                    mpi_u[i][j][0] += e[k][0] * mpi_f[i][j][k]
                    mpi_u[i][j][1] += e[k][1] * mpi_f[i][j][k]
                mpi_u[i][j][0] /= mpi_rho[i][j]
                mpi_u[i][j][1] /= mpi_rho[i][j]
        
        # 左右边界
        for j in range(1, LYv - 2):
            for k in range(Q):
                mpi_rho[0][j] = mpi_rho[1][j]
                mpi_f[0][j][k] = feq(k, mpi_rho[0][j], mpi_u[0][j]) + mpi_f[1][j][k] - feq(k, mpi_rho[1][j], mpi_u[1][j])
                mpi_rho[Lx_mpi - 1][j] = mpi_rho[Lx_mpi - 2][j]
                mpi_f[Lx_mpi - 1][j][k] = feq(k, mpi_rho[Lx_mpi - 1][j], mpi_u[Lx_mpi - 1][j]) + mpi_f[Lx_mpi - 2][j][k] - feq(k, mpi_rho[Lx_mpi - 2][j], mpi_u[Lx_mpi - 2][j])
        
        # 上边界（顶盖驱动流）
        for i in range(Lx_mpi):
            mpi_u[i][LYv - 2][0] = U
            mpi_rho[i][LYv - 2] = mpi_rho[i][LYv - 3]
            for k in range(Q):
                mpi_f[i][LYv - 2][k] = feq(k, mpi_rho[i][LYv - 2], mpi_u[i][LYv - 2]) + mpi_f[i][LYv - 3][k] - feq(k, mpi_rho[i][LYv - 3], mpi_u[i][LYv - 3])


@njit
def evolution(rho, u, f, u0):
    F = np.zeros((NY + 1, NX + 1, Q))
    for i in range(1, NX):
        for j in range(1, NY):
            for k in range(Q):
                ip = i - e[k][0]
                jp = j - e[k][1]
                F[i][j][k] = f[ip][jp][k] + (feq(k, rho[ip][jp], u[ip][jp]) - f[ip][jp][k]) / tau_f
    for i in range(1, NX):
        for j in range(1, NY):
            rho[i][j] = 0
            u0[i][j][0] = u[i][j][0]
            u0[i][j][1] = u[i][j][1]
            u[i][j][0] = 0
            u[i][j][1] = 0
            for k in range(Q):
                f[i][j][k] = F[i][j][k]
                rho[i][j] += f[i][j][k]
                u[i][j][0] += e[k][0] * f[i][j][k]
                u[i][j][1] += e[k][1] * f[i][j][k]
            u[i][j][0] /= rho[i][j]
            u[i][j][1] /= rho[i][j]
    for j in range(1, NY):
        for k in range(Q):
            rho[NX][j] = rho[NX - 1][j]
            f[NX][j][k] = feq(k, rho[NX - 1][j], u[NX][j]) + f[NX - 1][j][k] - feq(k, rho[NX - 1][j], u[NX - 1][j])
            rho[0][j] = rho[1][j]
            f[0][j][k] = feq(k, rho[1][j], u[0][j]) + f[1][j][k] - feq(k, rho[1][j], u[1][j])
    for i in range(NX + 1):
        for k in range(Q):
            rho[i][0] = rho[i][1]

            f[i][0][k] = feq(k, rho[i][0], u[i][0]) + f[i][1][k] - feq(k, rho[i][1], u[i][1])
            u[i][NY][0] = U
            rho[i][NY] = rho[i][NY - 1]
            f[i][NY][k] = feq(k, rho[i][NY], u[i][NY]) + f[i][NY - 1][k] - feq(k, rho[i][NY - 1], u[i][NY - 1])


def output(m, u):
    """输出流场数据"""
    if rank == 0:
        with open(f'cavity_{m}.dat', 'w') as out:
            out.write("Title=\"LBM Lid Driven Flow\"\n")
            out.write("VARIABLES=\"X\",\"Y\",\"U\",\"V\"\n")
            out.write(f"ZONE T=\"BOX\",I= {NX + 1},J= {NY + 1},F=POINT\n")
            for j in range(NY + 1):
                for i in range(NX + 1):
                    out.write(f"{i / Lx} {j / Ly} {u[i][j][0]} {u[i][j][1]}\n")



def writefileE1(Ub, A, T):
    """写入训练数据到Excel"""
    if rank == 0:
        filename = "endijtrain_011_1w.xlsx"
        
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
        
        rows_data = []
        for i in range(128):
            for j in range(128):
                # 根据PDF方法一，输入特征应该包括：
                # 1. |ω|: 旋度大小 = |A_xy - A_yx|
                # 2. θ: 散度 = A_xx + A_yy
                # 3. √(A_ij A_ij): 速度梯度张量模
                # 4. √(S_ij S_ij): 应变率张量模
                # 以及A_ij的各个分量
                omega_abs = np.abs(A[i][j][1] - A[i][j][2])  # 旋度大小 |ω|
                S_xy = 0.5 * (A[i][j][1] + A[i][j][2])  # 应变率张量的xy分量
                Sm = np.sqrt(A[i][j][0]*A[i][j][0] + S_xy*S_xy + S_xy*S_xy + A[i][j][3]*A[i][j][3])  # 应变率张量模
                
                row = [
                    float(i / 128),  # x坐标
                    float(j / 128),  # y坐标
                    omega_abs,  # |ω|: 旋度大小（输入特征0）
                    A[i][j][0] + A[i][j][3],  # θ: 散度（输入特征1）
                    np.sqrt(A[i][j][0]*A[i][j][0]+A[i][j][1]*A[i][j][1]+A[i][j][2]*A[i][j][2]+A[i][j][3]*A[i][j][3]),  # √(A_ij A_ij)（输入特征2）
                    Sm,  # √(S_ij S_ij)（输入特征3）
                    A[i][j][0],  # A_xx（输入特征4）
                    A[i][j][1],  # A_xy（输入特征5）
                    A[i][j][2],  # A_yx（输入特征6）
                    A[i][j][3],  # A_yy（输入特征7）
                    T[i][j][0],  # τ_xx（输出）
                    T[i][j][1],  # τ_xy（输出）
                    T[i][j][2],  # τ_yx（输出）
                    T[i][j][3]   # τ_yy（输出）
                ]
                rows_data.append(row)
        
        for row in rows_data:
            ws.append(row)
        
        wb.save(filename)

@njit
def error_mpi(mpi_u, mpi_u0, me_y, MY, Lx_mpi, LYv):
    """MPI版本的误差计算"""
    temp1 = 0.0
    temp2 = 0.0
    
    if me_y == 0:
        # 下边界进程：从j=2开始（跳过虚格点和物理边界）
        for i in range(1, Lx_mpi - 1):
            for j in range(2, LYv - 1):
                temp1 += ((mpi_u[i][j][0] - mpi_u0[i][j][0]) ** 2 + (mpi_u[i][j][1] - mpi_u0[i][j][1]) ** 2)
                temp2 += (mpi_u[i][j][0] ** 2 + mpi_u[i][j][1] ** 2)
    elif me_y == MY:
        # 上边界进程：到LYv-2（跳过虚格点）
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 2):
                temp1 += ((mpi_u[i][j][0] - mpi_u0[i][j][0]) ** 2 + (mpi_u[i][j][1] - mpi_u0[i][j][1]) ** 2)
                temp2 += (mpi_u[i][j][0] ** 2 + mpi_u[i][j][1] ** 2)
    else:
        # 内部进程
        for i in range(1, Lx_mpi - 1):
            for j in range(1, LYv - 1):
                temp1 += ((mpi_u[i][j][0] - mpi_u0[i][j][0]) ** 2 + (mpi_u[i][j][1] - mpi_u0[i][j][1]) ** 2)
                temp2 += (mpi_u[i][j][0] ** 2 + mpi_u[i][j][1] ** 2)
    
    return temp1, temp2


def mpi_gather_u(mpi_u, u_global, me_y, MY, Lx_mpi, LYv, Ly_mpi):
    """收集所有进程的速度数据到rank 0"""
    # 提取每个进程的有效数据（排除虚格点）
    if me_y == 0:
        local_u = mpi_u[:, 1:LYv-1, :].copy()  # 排除下虚格点
    elif me_y == MY:
        local_u = mpi_u[:, 1:LYv-2, :].copy()  # 排除上下虚格点
    else:
        local_u = mpi_u[:, 1:LYv-1, :].copy()  # 排除上下虚格点
    
    # 收集所有进程的数据
    gathered_data = comm.gather(local_u, root=0)
    
    if rank == 0:
        # 合并所有进程的数据
        start_y = 0
        for data in gathered_data:
            end_y = start_y + data.shape[1]
            u_global[:, start_y:end_y, :] = data
            start_y = end_y
        return u_global
    else:
        return None


def main():
    if rank == 0:
        print(f"MPI并行计算：使用 {size} 个进程")
        print("=" * 60)
    
    # 初始化
    me_y, MY = init_mpi(mpi_f, mpi_u, mpi_rho, rank, size, Ly_mpi, LYv, Lx_mpi)
    
    # 全局数组（只在rank 0）
    T = np.zeros((128, 128, 4))  # 记录亚格子应力张量
    Ub = np.zeros((128, 128, 2))  # 记录滤波后的粗格子速度
    A = np.zeros((128, 128, 4))  # 记录速度梯度张量
    Am = np.zeros((128, 128, 1))  # 记录速度梯度张量的模
    
    n = 0
    start_time = time.time()
    
    while True:
        # MPI通信：发送和接收边界数据
        befor_send(allsend, mpi_f, mpi_u, mpi_rho, Lx_mpi, Q, LYv)
        mpi_sendarecv(me_y, MY)
        after_receive(mpi_f, mpi_u, mpi_rho, allReceive, Lx_mpi, Q, LYv)
        
        # LBM演化
        mpi_evolution(mpi_f, mpi_u, mpi_rho, mpi_F, mpi_u0, me_y, MY, Lx_mpi, LYv)
        
        if n % 10000 == 0:
            # 计算误差
            temp1, temp2 = error_mpi(mpi_u, mpi_u0, me_y, MY, Lx_mpi, LYv)
            temp1_all = comm.reduce(temp1, op=MPI.SUM, root=0)
            temp2_all = comm.reduce(temp2, op=MPI.SUM, root=0)
            
            if rank == 0:
                err = np.sqrt(temp1_all) / (np.sqrt(temp2_all) + 1e-30)
                elapsed = time.time() - start_time
                iter_per_sec = n / elapsed if elapsed > 0 else 0
                print(f"The {n}th computation result:")
                print(f"The max relative error of uv is: {err:.6e}")
                print(f"Speed: {iter_per_sec:.1f} iterations/sec, Elapsed: {elapsed:.1f}s")
                print("-" * 60)
            
            if n >= 10000:
                # 收集完整速度场用于后处理
                u_collected = mpi_gather_u(mpi_u, u_global, me_y, MY, Lx_mpi, LYv, Ly_mpi)
                
                if rank == 0:
                    # 计算滤波和梯度（需要完整速度场）
                    tforce(T, u_collected, Ub)
                    ronudU(Ub, A, Am)
                    
                    if n % 20000 == 0:
                        output(n, u_collected)
                    if n % 5000 == 0:
                        writefileE1(Ub, A, T)
                    
                    if n == 500000:
                        break
                    if err < 1.0e-6:
                        break
        
        n += 1
    
    if rank == 0:
        total_time = time.time() - start_time
        print("=" * 60)
        print(f"Simulation completed!")
        print(f"Total iterations: {n}")
        print(f"Total time: {total_time:.2f} seconds")
        print(f"Average speed: {n/total_time:.1f} iterations/sec")


if __name__ == "__main__":
    main()

