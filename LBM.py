import numpy as np
from numba import njit
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

# 创建一个空的DataFrame
df = pd.DataFrame()

# 将DataFrame写入Excel文件
df.to_excel('.train.xlsx', index=False)
wb = openpyxl.Workbook()

# 获取默认创建的工作表
ws = wb.active

# 将工作表重命名为"Data"
ws.title = "Data"
Q = 9
NX = 1023
NY = 1023
U = 0.1

e = np.array([[0, 0], [1, 0], [0, 1], [-1, 0], [0, -1], [1, 1], [-1, 1], [-1, -1], [1, -1]])
w = np.array([4.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 9, 1.0 / 36, 1.0 / 36, 1.0 / 36, 1.0 / 36])

dx = 1.0
dy = 1.0
Lx = dx * NY
Ly = dy * NX
dt = dx
c = dx / dt
rho0 = 1.0
Re = 10000
niu = U * Lx / Re
tau_f = 3.0 * niu + 0.5


@njit
def tforce(T,u,Ub):
    for i in range(1,127):
        for j in range(1,127):
               ujj=0.0
               uii = 0.0
               uij=0.0
               ui=0.0
               uj=0.0
               for i1 in range(8):
                   for j1 in range(8):
                      if i1==1 or i1==8 or j1==1 or j1==8:
                          ui+=u[i*8+i1][j*8+j1][0]*0.5
                          uj += u[i * 8 + i1][j * 8 + j1][1]*0.5
                          uij+=u[i*8+i1][j*8+j1][0]*u[i*8+i1][j*8+j1][1]*0.5
                          uii+=u[i*8+i1][j*8+j1][0]*u[i*8+i1][j*8+j1][0]*0.5
                          ujj += u[i * 8 + i1][j * 8 + j1][1] * u[i * 8 + i1][j * 8 + j1][1] * 0.5
                      else :
                           ui+=u[i*8+i1][j*8+j1][0]
                           uj+=u[i*8+i1][j*8+j1][1]
                           uij+=u[i*8+i1][j*8+j1][0]*u[i*8+i1][j*8+j1][1]
                           uii += u[i * 8 + i1][j * 8 + j1][0] * u[i * 8 + i1][j * 8 + j1][0]
                           ujj += u[i * 8 + i1][j * 8 + j1][1] * u[i * 8 + i1][j * 8 + j1][1]
               T[i][j][0]=ui/64.0*ui/64.0-uii/64.0   #txx
               T[i][j][1]=ui/64.0*uj/64.0-uij/64.0   #txy
               T[i][j][2]=ui/64.0*uj/64.0-uij/64.0   #tyx
               T[i][j][3] =uj/64.0*uj/64.0-ujj/64.0  #tyy
               Ub[i][j][0] = ui / 64.0               #ux~
               Ub[i][j][1] = uj / 64.0               #uy~

@njit
def ronudU(Ub,A,Am):
    #梯度张量
    for i in range(1,127):
        for j in range(1,127):
            if i==127 or j==127:
                A[i][j][0] = Ub[i][j][0] - Ub[i][j-1][0]  # Axx
                A[i][j][1] = Ub[i][j][1] - Ub[i][j-1][1]  # Axy
                A[i][j][2] = Ub[i][j][0] - Ub[i-1][j][0]  # Ayx
                A[i][j][3] = Ub[i][j][1] - Ub[i-1][j][1]  # Ayy
            A[i][j][0] = Ub[i][j + 1][0] - Ub[i][j][0]  # Axx
            A[i][j][1] = Ub[i][j + 1][1] - Ub[i][j][1]  # Axy
            A[i][j][2] = Ub[i + 1][j][0] - Ub[i][j][0]  # Ayx
            A[i][j][3] = Ub[i + 1][j][1] - Ub[i][j][1]  # Ayy
            Am[i][j]=np.sqrt(A[i][j][0]*A[i][j][0]+A[i][j][1]*A[i][j][1]+A[i][j][2]*A[i][j][2]+A[i][j][3]*A[i][j][3])



@njit
def feq(k, rho, u):
    eu = e[k][0] * u[0] + e[k][1] * u[1]
    uv = u[0] * u[0] + u[1] * u[1]
    return w[k] * rho * (1.0 + 3.0 * eu + 4.5 * eu ** 2 - 1.5 * uv)


@njit
def evolution(rho, u, f, u0):
    F = np.zeros((NY + 1, NX + 1, Q))
    for i in range(1, NX):
        for j in range(1, NY):
            for k in range(Q):
                ip = i - e[k][0]
                jp = j - e[k][1]
                F[i][j][k] = f[ip][jp][k] + (feq(k, rho[ip][jp], u[ip][jp]) - f[ip][jp][k]) / tau_f
    for i in range(1, NX):
        for j in range(1, NY):
            rho[i][j] = 0
            u0[i][j][0] = u[i][j][0]
            u0[i][j][1] = u[i][j][1]
            u[i][j][0] = 0
            u[i][j][1] = 0
            for k in range(Q):
                f[i][j][k] = F[i][j][k]
                rho[i][j] += f[i][j][k]
                u[i][j][0] += e[k][0] * f[i][j][k]
                u[i][j][1] += e[k][1] * f[i][j][k]
            u[i][j][0] /= rho[i][j]
            u[i][j][1] /= rho[i][j]
    for j in range(1, NY):
        for k in range(Q):
            rho[NX][j] = rho[NX - 1][j]
            f[NX][j][k] = feq(k, rho[NX - 1][j], u[NX][j]) + f[NX - 1][j][k] - feq(k, rho[NX - 1][j], u[NX - 1][j])
            rho[0][j] = rho[1][j]
            f[0][j][k] = feq(k, rho[1][j], u[0][j]) + f[1][j][k] - feq(k, rho[1][j], u[1][j])
    for i in range(NX + 1):
        for k in range(Q):
            rho[i][0] = rho[i][1]

            f[i][0][k] = feq(k, rho[i][0], u[i][0]) + f[i][1][k] - feq(k, rho[i][1], u[i][1])
            u[i][NY][0] = U
            rho[i][NY] = rho[i][NY - 1]
            f[i][NY][k] = feq(k, rho[i][NY], u[i][NY]) + f[i][NY - 1][k] - feq(k, rho[i][NY - 1], u[i][NY - 1])


def output(m, u):
    with open(f'cavity_{m}.dat', 'w') as out:
        out.write("Title=\"LBM Lid Driven Flow\"\n")
        out.write("VARIABLES=\"X\",\"Y\",\"U\",\"V\"\n")
        out.write(f"ZONE T=\"BOX\",I= {NX + 1},J= {NY + 1},F=POINT\n")
        for j in range(NY + 1):
            for i in range(NX + 1):
                out.write(f"{i / Lx} {j / Ly} {u[i][j][0]} {u[i][j][1]}\n")



def writefileE1(Ub,A,T):
    filename = "endijtrain_011_1w.xlsx"

    try:
        # 尝试打开现有的Excel文件
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # 如果文件不存在，创建新的Excel文件并添加一个名为Data的工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

    max_row = ws.max_row if ws.max_row > 0 else 1

    for i in range(128):
        for j in range(128):
            row = [
                        #涡度，散度，应力梯度张量
                float(i / 128),
                float(j / 128),
                 np.sqrt(Ub[i][j][0]*Ub[i][j][0]+Ub[i][j][1]*Ub[i][j][1]),A[i][j][0]+A[i][j][3],
                 np.sqrt(A[i][j][0]*A[i][j][0]+A[i][j][1]*A[i][j][1]+A[i][j][2]*A[i][j][2]+A[i][j][3]*A[i][j][3]),
                 np.sqrt(A[i][j][0]*A[i][j][0]+(A[i][j][1]+A[i][j][2])*0.5*(A[i][j][1]+A[i][j][2])*0.5*2+A[i][j][3]*A[i][j][3])
                ,A[i][j][0],A[i][j][1],A[i][j][2],A[i][j][3],T[i][j][0],T[i][j][1],T[i][j][2],T[i][j][3]

            ]
            ws.append(row)

    # 保存Excel文件
    wb.save(filename)

@njit
def error(u, u0):
    temp1 = 0
    temp2 = 0
    for i in range(1, NY):
        for j in range(1, NX):
            temp1 += ((u[i][j][0] - u0[i][j][0]) ** 2 + (u[i][j][1] - u0[i][j][1]) ** 2)
            temp2 += (u[i][j][0] ** 2 + u[i][j][1] ** 2)
    temp1 = np.sqrt(temp1)
    temp2 = np.sqrt(temp2)
    return temp1 / (temp2 + 1e-30)


def main():
    rho = np.ones((NY + 1, NX + 1))
    u = np.zeros((NY + 1, NX + 1, 2))
    f = np.zeros((NY + 1, NX + 1, Q))
    u0 = np.zeros((NY + 1, NX + 1, 2))  # 将u0初始化为全零数组
    T = np.zeros((128, 128,4))  # 记录亚格子应力张量
    Ub=np.zeros((128, 128,2))   #记录滤波后的粗格子速度
    A = np.zeros((128, 128, 4)) #记录速度梯度张量
    Am = np.zeros((128, 128, 1))  # 记录速度梯度张量的模

    n = 0
    while True:
        evolution(rho, u, f, u0)
        if n % 10000 == 0:
            err = error(u, u0)
            tforce(T, u,Ub)
            ronudU(Ub,A,Am)
            print(f"The {n}th computation result:")
            print(f"The u,v pf point (NX/2,NY/2) is: {u[NY // 2][NX // 2][0]}, {u[NY // 2][NX // 2][1]}")
            print(f"The max relative error of uv is: {err:.6e}")
            if n >= 10000:
                if n % 20000 == 0:
                    output(n, u)
                if n% 5000==0:
                    writefileE1(Ub,A,T)
                if n==500000:
                    break
                if err < 1.0e-6:

                    break
        n += 1


if __name__ == "__main__":
    main()